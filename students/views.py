from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import FileResponse
from openpyxl import load_workbook
import hashlib
import os
from openpyxl import Workbook
from django.http import HttpResponse

from .models import Student, UploadHistory
from .tasks import send_welcome_email
from .utils import create_remark_file   # adjust import if needed


def upload_excel(request):

    if request.method == "POST":

        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please select an Excel file.")
            return redirect("upload_excel")

        try:

            # Read file content for hash
            file_content = excel_file.read()
            file_hash = hashlib.sha256(file_content).hexdigest()

            # Check duplicate file upload
            if UploadHistory.objects.filter(file_hash=file_hash).exists():
                messages.error(
                    request,
                    "This file has already been uploaded."
                )
                return redirect("upload_excel")

            # Reset pointer
            excel_file.seek(0)

            workbook = load_workbook(
                excel_file,
                data_only=True
            )

            worksheet = workbook.active

            uploaded_count = 0
            duplicate_count = 0
            invalid_count = 0

            remarks = []

            # Read rows
            for row in worksheet.iter_rows(
                min_row=2,
                values_only=True
            ):

                name, age, email = row

                # Validation
                if not name or not age or not email:

                    invalid_count += 1

                    remarks.append({
                        "name": name or "",
                        "age": age or "",
                        "email": email or "",
                        "remark": "Invalid data - Missing field"
                    })

                    continue

                # Check duplicate student
                if Student.objects.filter(email=email).exists():

                    duplicate_count += 1

                    remarks.append({
                        "name": name,
                        "age": age,
                        "email": email,
                        "remark": "Already Exists"
                    })

                else:

                    student = Student.objects.create(
                        name=name,
                        age=age,
                        email=email
                    )

                    uploaded_count += 1

                    # Send email in background
                    send_welcome_email.delay(student.id)

                    remarks.append({
                        "name": name,
                        "age": age,
                        "email": email,
                        "remark": "Uploaded Successfully"
                    })

            # Create remark file
            remark_file_path = None

            if remarks:
                remark_file_path = create_remark_file(remarks)

            # Save upload history
            upload_history = UploadHistory.objects.create(
                file_name=excel_file.name,
                file_hash=file_hash,
            )

            if remark_file_path:
                upload_history.remark_file = (
                    "remarks/" + os.path.basename(remark_file_path)
                )

            upload_history.save()

            messages.success(
                request,
                f"""
Upload Completed.

New Records: {uploaded_count}

Duplicate Records: {duplicate_count}

Invalid Records: {invalid_count}
"""
            )

            # Automatically download remark file
            if remark_file_path and os.path.exists(remark_file_path):

                return FileResponse(
                    open(remark_file_path, "rb"),
                    as_attachment=True,
                    filename=os.path.basename(remark_file_path),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            return redirect("user_list")

        except Exception as e:

            messages.error(
                request,
                f"Error processing the file: {str(e)}"
            )

            return redirect("upload_excel")

    # IMPORTANT: Handle GET request
    return render(request, "upload.html")



def user_list(request):

    students = Student.objects.all()


    return render(

        request,

        "user_list.html",

        {
            "students": students
        }

    )




def download_excel(request):


    workbook = Workbook()


    worksheet = workbook.active


    worksheet.title = "Students"



    worksheet.append([

        "Name",

        "Age",

        "Email"

    ])



    students = Student.objects.all()



    for student in students:


        worksheet.append([

            student.name,

            student.age,

            student.email

        ])




    response = HttpResponse(

        content_type=

        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )



    response["Content-Disposition"] = (

        'attachment; filename="students.xlsx"'

    )



    workbook.save(response)



    return response


def upload_history(request):

    history = UploadHistory.objects.all().order_by(
        "-uploaded_at"
    )

    return render(
        request,
        "upload_history.html",
        {
            "history": history
        }
    )